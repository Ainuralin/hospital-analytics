import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sqlalchemy import create_engine
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
import plotly.express as px
from config import DB_USER, DB_PASSWORD, DB_HOST, DB_PORT, DB_NAME

engine = create_engine(f'postgresql://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}')

os.makedirs("charts", exist_ok=True)

def save_and_report(df, title, filename, fig=None):
    if df.empty:
        print(f"⚠ {title}: данных нет, график не построен.")
        return
    if fig is None:
        fig = plt.gcf()
    filepath = os.path.join("charts", filename)
    fig.savefig(filepath, dpi=300, bbox_inches="tight")
    print(f"▶ {title}: {len(df)} строк(и) получено. График сохранён -> {filepath}")
    plt.close(fig)

def remove_tz(df):
    for col in df.columns:
        if pd.api.types.is_datetime64_any_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)
    return df

# 1. Surgery Bar Chart
df_surgery = pd.read_sql("""
SELECT surgery_type, COUNT(surgery_id) AS surgery_count
FROM surgeryrecord
GROUP BY surgery_type;
""", engine)

fig, ax = plt.subplots(figsize=(12, 7))
bars = ax.bar(df_surgery["surgery_type"], df_surgery["surgery_count"],
                width=0.6, color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd'])
ax.bar_label(bars, padding=3)
ax.set_title("Количество операций по типам", fontsize=16, pad=20)
ax.set_xlabel("Тип операции", fontsize=13)
ax.set_ylabel("Количество операций", fontsize=13)
plt.xticks(rotation=45, ha="right", fontsize=11)
plt.tight_layout()
save_and_report(df_surgery, "Bar chart (операции по типам)", "bar_surgery.png", fig)

# 2. Horizontal Bar – Врачи по отделениям
df_doctors = pd.read_sql("""
SELECT d.dept_name, COUNT(doc.doct_id) AS doctor_count
FROM doctor doc
JOIN department d ON doc.dept_id = d.dept_id
GROUP BY d.dept_name
ORDER BY doctor_count;
""", engine)

fig, ax = plt.subplots(figsize=(10, 6))
bars = ax.barh(df_doctors["dept_name"], df_doctors["doctor_count"],
                color="skyblue", edgecolor="navy", linewidth=1)
for i, v in enumerate(df_doctors["doctor_count"].values):
    ax.text(v + 0.1, i, str(v), va="center", fontweight="bold")
ax.set_title("Количество врачей по отделениям", fontsize=16, pad=20)
ax.set_xlabel("Количество врачей", fontsize=12)
ax.set_ylabel("Отделение", fontsize=12)
ax.grid(axis="x", alpha=0.3)
plt.tight_layout()
save_and_report(df_doctors, "Horizontal bar (врачи по отделениям)", "barh_doctors.png", fig)

# 3. Histogram – Возраст пациентов
df_patients = pd.read_sql("SELECT patient_id, date_of_birth FROM patients", engine)
df_patients['age'] = (pd.to_datetime('today') - pd.to_datetime(df_patients['date_of_birth'])).dt.days // 365

fig, ax = plt.subplots(figsize=(10, 6))
ax.hist(df_patients['age'], bins=20, color='skyblue', alpha=0.7, edgecolor='black')
mean_age = df_patients['age'].mean()
ax.axvline(mean_age, color='red', linestyle='--', linewidth=2, label=f'Mean: {mean_age:.1f}')
ax.set_title('Age Distribution of Patients', fontsize=16, pad=15)
ax.set_xlabel('Age', fontsize=12)
ax.set_ylabel('Number of Patients', fontsize=12)
ax.legend()
ax.grid(axis='y', alpha=0.3)
plt.tight_layout()
save_and_report(df_patients, "Histogram (возраст пациентов)", "hist_age_patients.png", fig)

# 4. Line Chart – Пациенты по отделениям по месяцам
df_line = pd.read_sql("""
SELECT 
    d.dept_name,
    DATE_TRUNC('month', a.appointment_date) AS month,
    COUNT(a.appointment_id) AS patient_count
FROM appointment a
JOIN doctor doc ON a.doct_id = doc.doct_id
JOIN department d ON doc.dept_id = d.dept_id
GROUP BY d.dept_name, DATE_TRUNC('month', a.appointment_date)
ORDER BY month;
""", engine)

df_line = remove_tz(df_line)
df_pivot = df_line.pivot(index="month", columns="dept_name", values="patient_count").fillna(0)

fig, ax = plt.subplots(figsize=(14, 8))
for dept in df_pivot.columns:
    ax.plot(df_pivot.index, df_pivot[dept], label=f"{dept} (fact)", linewidth=2, marker='o', markersize=4)
    rolling_avg = df_pivot[dept].rolling(window=3).mean()
    ax.plot(rolling_avg.index, rolling_avg, '--', alpha=0.7, linewidth=1, label=f"{dept} (trend)")
ax.set_title("Количество пациентов по отделениям (с трендами)", fontsize=16, pad=20)
ax.set_xlabel("Дата", fontsize=12)
ax.set_ylabel("Количество пациентов", fontsize=12)
ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
ax.grid(True, alpha=0.3)
plt.xticks(rotation=45)
plt.tight_layout()
save_and_report(df_line, "Line chart (пациенты по отделениям по месяцам)", "line_departments.png", fig)

# 5. Pie Chart – Распределение пациентов
df_pie = pd.read_sql("""
SELECT d.dept_name, COUNT(p.patient_id) AS patient_count
FROM patients p
JOIN appointment a ON p.patient_id = a.patient_id
JOIN doctor doc ON a.doct_id = doc.doct_id
JOIN department d ON doc.dept_id = d.dept_id
GROUP BY d.dept_name;
""", engine)

if not df_pie.empty:
    fig, ax = plt.subplots(figsize=(10, 8))
    colors = ['#FF9999', '#66B2FF', '#99FF99', '#FFCC99', '#FFD700']
    wedges, texts, autotexts = ax.pie(df_pie["patient_count"],
                                    labels=df_pie["dept_name"],
                                    autopct='%1.1f%%',
                                    colors=colors[:len(df_pie)],
                                    explode=[0.05] + [0]*(len(df_pie)-1),
                                    shadow=True,
                                    startangle=90)
    for autotext in autotexts:
        autotext.set_color('white')
        autotext.set_weight('bold')
    ax.set_title("Распределение пациентов по отделениям", fontsize=16, pad=20)
    plt.tight_layout()
    save_and_report(df_pie, "Pie chart (пациенты по отделениям)", "pie_departments.png", fig)

# 6. Scatter Plot – Возраст vs Приёмы
df_scatter = pd.read_sql("""
SELECT 
    p.patient_id,
    EXTRACT(YEAR FROM AGE(p.date_of_birth)) AS age,
    d.dept_name,
    COUNT(a.appointment_id) AS num_appointments,
    SUM(a.payment_amount) AS total_payment
FROM patients p
JOIN appointment a ON p.patient_id = a.patient_id
JOIN doctor doc ON a.doct_id = doc.doct_id
JOIN department d ON doc.dept_id = d.dept_id
GROUP BY p.patient_id, age, d.dept_name
ORDER BY age
""", engine)

fig, ax = plt.subplots(figsize=(12, 8))
departments = df_scatter['dept_name'].unique()
colors = plt.cm.tab10.colors[:len(departments)]

for i, dept in enumerate(departments):
    dept_data = df_scatter[df_scatter['dept_name'] == dept]
    
    # X немного разбросан
    x_jitter = dept_data['age'] + np.random.uniform(-0.3, 0.3, size=len(dept_data))
    
    # Y только чётные целые
    y_jitter = 2 * (dept_data['num_appointments'].astype(int) // 2)

    ax.scatter(
        x_jitter, 
        y_jitter,
        s=np.clip(dept_data['total_payment'] / 50, 20, 200),
        alpha=0.6, 
        c=[colors[i]]*len(dept_data),
        label=dept, 
        edgecolors='black', 
        linewidth=0.5
    )

ax.set_title("Patients by Age and Number of Appointments", fontsize=16, pad=20)
ax.set_xlabel("Age", fontsize=12)
ax.set_ylabel("Number of Appointments", fontsize=12)
ax.grid(True, alpha=0.3)
ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

# Убираем дробные метки на оси Y
y_ticks = range(0, int(df_scatter['num_appointments'].max()) + 2, 2)
ax.set_yticks(y_ticks)

plt.tight_layout()
save_and_report(df_scatter, "Scatter plot (возраст vs приёмы)", "scatter_patients.png", fig)


# 7. Экспорт всех данных в Excel
def export_to_excel(dataframes_dict, filename):
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df = remove_tz(df)
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, startcol=1)
            ws = writer.sheets[sheet_name]
            ws['B1'] = f'Report: {sheet_name}'
            ws['B1'].font = Font(size=16, bold=True, color='1F4E79')
            ws.freeze_panes = "B3"

            header_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            header_font = Font(bold=True, color='1F4E79')
            for col_num in range(2, len(df.columns) + 2):
                cell = ws.cell(row=2, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for column in ws.columns:
                max_length = max((len(str(cell.value)) if cell.value is not None else 0 for cell in column))
                ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

            numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
            for i, col_name in enumerate(df.columns, start=2):
                if col_name in numeric_columns and len(df) >= 1:
                    col_letter = ws.cell(row=1, column=i).column_letter
                    range_address = f"{col_letter}3:{col_letter}{len(df)+2}"
                    color_scale_rule = ColorScaleRule(
                        start_type="min", start_color="63BE7B",
                        mid_type="percentile", mid_value=50, mid_color="FFDD71",
                        end_type="max", end_color="F8696B"
                    )
                    ws.conditional_formatting.add(range_address, color_scale_rule)

            ws.auto_filter.ref = f"B2:{ws.cell(row=len(df)+2, column=len(df.columns)+1).coordinate}"

            thin_border = Border(left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            for row in range(2, len(df)+3):
                for col in range(2, len(df.columns)+2):
                    ws.cell(row=row, column=col).border = thin_border

    print(f"✅ File '{filename}' successfully created with {len(dataframes_dict)} sheets.")

dataframes = {
    "Surgery": df_surgery,
    "Doctors": df_doctors,
    "Patients Age": df_patients,
    "Patients per Dept": df_line,
    "Patient Distribution": df_pie,
    "Appointments Scatter": df_scatter
}

export_to_excel(dataframes, "exports/hospital_report.xlsx")

# 8. Interactive Plotly – Поступления по кроватям
def interactive_bed_plot(engine):
    query = """
    SELECT br.admission_id,
            br.bed_no,
            br.patient_id,
            br.nurse_id,
            br.helper_id,
            br.admission_date,
            br.discharge_date,
            br.amount,
            br.mode_of_payment,
            p.fname || ' ' || p.lname AS patient_name
    FROM bedrecords br
    JOIN patients p ON br.patient_id = p.patient_id
    ORDER BY br.admission_date
    """
    df = pd.read_sql(query, engine)
    df['admission_date'] = pd.to_datetime(df['admission_date'], errors='coerce')
    df = df.dropna(subset=['admission_date'])
    df['month'] = df['admission_date'].dt.to_period('M').astype(str)

    df_agg = df.groupby(['month', 'bed_no', 'mode_of_payment']).agg(
        total_amount=pd.NamedAgg(column='amount', aggfunc='sum'),
        patient_count=pd.NamedAgg(column='patient_id', aggfunc='count')
    ).reset_index()

    fig = px.scatter(
        df_agg,
        x='bed_no',
        y='total_amount',
        size='patient_count',
        color='mode_of_payment',
        animation_frame='month',
        hover_data={'bed_no': True, 'total_amount': True, 'patient_count': True, 'mode_of_payment': True},
        title='Поступления пациентов по кроватям и оплате'
    )

    fig.update_layout(
        xaxis_title='Номер кровати',
        yaxis_title='Сумма оплат',
        legend_title='Способ оплаты',
        template='plotly_white'
    )

    fig.show()

interactive_bed_plot(engine)
